import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from openpyxl.workbook.workbook import Workbook
import streamlit as st
import requests
from streamlit_lottie import st_lottie
from openpyxl.styles import Font, Border, Side
import time


col3,col4,col8= st.columns(3)
with col3:
    st.header('Delta...')
with col4:
    selected = st.radio("Select model type",options=['Full model','Earnings'])
with col8:
    R_name = st.text_input("Reviewer Name")

#function
# For fetching all SRC from hyerlinks

def extract_hyperlinks_from_excel(excel_file):
    # Create a dictionary to store the hyperlinks.
    #src_num_dict = {}
    list_of_dict = {}

    # Load the Excel file.
    workbook = load_workbook(excel_file, data_only=True)
    
    # Iterate through all sheets in the workbook.
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Iterate through all cells in the sheet.
        for row in sheet.iter_rows():
            for cell in row:
                # Check if the cell contains a hyperlink.
                if cell.hyperlink is not None:
                    src_num_dict = {}
                    cell_location = cell.coordinate  # Cell location like 'A1'
                    src_num = cell.hyperlink.target.split("/")[-1]  # The hyperlink URL
                    src_num_dict[src_num] = [src_num, cell_location,cell.column,cell.value,cell]
    
                    list_of_dict[src_num]=src_num_dict

    return list_of_dict

# For fetching all SRC from hyerlinks
def extract_hyperlinks_from_excel_earnings(excel_file,column_earning):
    # Create a dictionary to store the hyperlinks.
    #src_num_dict = {}
    list_of_dict = {}

    # Load the Excel file.
    workbook = load_workbook(excel_file, data_only=True)

    # Iterate through all sheets in the workbook.
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Iterate through all cells in the sheet.
        for row in sheet.iter_rows():
            for cell in row[column_earning:]:
                # Check if the cell contains a hyperlink.
                if cell.hyperlink is not None:
                    src_num_dict = {}
                    lable_name = sheet.cell(cell.row,1).value
                    unit=sheet.cell(cell.row,3).value
                    period= sheet.cell(cell.row,4).value
                    cell_location = cell.coordinate  # Cell location like 'A1'
                    src_num = cell.hyperlink.target.split("/")[-1]  # The hyperlink URL
                    src_num_dict[src_num] = [src_num, cell_location,cell.column,cell.value,cell,lable_name,unit,period]

                    list_of_dict[src_num]=src_num_dict

    return list_of_dict

#SRC for units and periods
def unit_period_dict(wb_fn,deleted_src,data_added_src,AR_src):
    unit_and_period_data ={}
    for row in range (1,(wb_fn.max_row+1)):
        cell_unit = wb_fn.cell(row = row, column= 3)
        cell_period = wb_fn.cell(row = row, column= 4)
        latest_cell = wb_fn.cell(row = row, column= wb_fn.max_column)
        if cell_unit.value is not None:
            if latest_cell.value is None:
                for column in range ((-1*wb_fn.max_column),4):
                    if column < 0:
                        iter_cell = wb_fn.cell(row = row, column= (-1*column))
                    elif column > 0:
                        iter_cell = wb_fn.cell(row = row, column= (column))
                        #print(iter_cell.coordinate)
                    if iter_cell.hyperlink is not None:
                        if (iter_cell.hyperlink.target.split("/")[-1]) in deleted_src:
                        # print(iter_cell.coordinate)
                            pass
                        elif (iter_cell.hyperlink.target.split("/")[-1]) in data_added_src:
                            pass
                        
                        elif (iter_cell.hyperlink.target.split("/")[-1]) in AR_src:
                            
                            unit = cell_unit.value
                            period = cell_period.value
                            src_num_dict = {}
                            cell_location = iter_cell.coordinate  # Cell location like 'A1'
                            src_num = iter_cell.hyperlink.target.split("/")[-1]  # The hyperlink URL
                            src_num_dict[src_num] = [unit,period, cell_location,iter_cell]
                            unit_and_period_data[src_num]=src_num_dict
                            break

            elif latest_cell.hyperlink is not None:
                for column in range ((-1*wb_fn.max_column),4):
                        if column < 0:
                            iter_cell = wb_fn.cell(row = row, column= (-1*column))
                        elif column > 0:
                            iter_cell = wb_fn.cell(row = row, column= (column))
                            #print(iter_cell.coordinate)
                        #print(iter_cell.coordinate)
                        #print(iter_cell.value)
                        if iter_cell.hyperlink is not None:
                            if (iter_cell.hyperlink.target.split("/")[-1]) in deleted_src:
                                pass
                            
                            elif (iter_cell.hyperlink.target.split("/")[-1]) in data_added_src:
                                pass
                            
                            elif (iter_cell.hyperlink.target.split("/")[-1]) in AR_src:
                                unit = cell_unit.value
                                period = cell_period.value
                                src_num_dict = {}
                                cell_location = iter_cell.coordinate  # Cell location like 'A1'
                                src_num = iter_cell.hyperlink.target.split("/")[-1]  # The hyperlink URL
                                src_num_dict[src_num] = [unit,period, cell_location,iter_cell]
                                unit_and_period_data[src_num]=src_num_dict
                                break

    return unit_and_period_data

#merging functions
def merge_unmerg_dict(wb_fn):
    merge_unmerg ={}
    for row in range (1,(wb_fn.max_row+1)):
        cell_unit = wb_fn.cell(row = row, column= 3)
        latest_cell = wb_fn.cell(row = row, column= wb_fn.max_column)
        if cell_unit.value is not None:
            if latest_cell.value is None:
                for column in range ((-1*wb_fn.max_column),4):
                    if column < 0:
                        iter_cell = wb_fn.cell(row = row, column= (-1*column))
                    elif column > 0:
                        iter_cell = wb_fn.cell(row = row, column= (column))
                    if iter_cell.hyperlink is not None:
                        src_num_dict = {}
                        cell_location = iter_cell.coordinate  # Cell location like 'A1'
                        src_num = iter_cell.hyperlink.target.split("/")[-1]  # The hyperlink URL
                        src_num_dict[src_num] = [iter_cell.row, cell_location]
                        merge_unmerg[src_num]=src_num_dict
                        break

            elif latest_cell.hyperlink is not None:
                src_num_dict = {}
                cell_location = latest_cell.coordinate  # Cell location like 'A1'
                src_num = latest_cell.hyperlink.target.split("/")[-1]  # The hyperlink URL
                src_num_dict[src_num] = [latest_cell.row, cell_location]
                merge_unmerg[src_num]=src_num_dict
    return merge_unmerg

def All_SRC_in_ROW(wb_fn,row,data_added_src,deleted_src):
    columns = wb_fn.max_column
    cell_unit = wb_fn.cell(row = row, column= 3)
    src_list =[]
    ar_row_list = []
    fr_row_list = []
    for column in range (4,(columns+1)):
        iter_cell = wb_fn.cell(row = row, column= column)
        if iter_cell.hyperlink is not None:
            src_num = iter_cell.hyperlink.target.split("/")[-1]
            if src_num in data_added_src:
                ar_row_list.append(src_num)
                
            elif src_num in deleted_src:
                fr_row_list.append(src_num)
                
            else:
                src_list.append(src_num)
                ar_row_list.append(src_num)
                fr_row_list.append(src_num)
        else:
            pass
    return src_list,ar_row_list,fr_row_list


def load_lottiurl(url: str):
        r = requests.get(url)
        if r.status_code != 200:
            return None
        
        return r.json()
#GIF loading

loader = load_lottiurl('https://lottie.host/289ca56b-6dbb-4337-b488-895f72a1c7cb/FpIA3aCqcm.json')
done_gif = load_lottiurl('https://lottie.host/43869007-4076-48ce-8d31-c9298325d54d/4JouEu0HdT.json')
error_gif = load_lottiurl('https://lottie.host/872f9d6e-08cb-4beb-831e-2a03ae581c90/AtWtDSyzhh.json')


if selected == "Full model":
    if R_name:
        col1,col2 = st.columns(2)
        with col1:
            AR = st.file_uploader("Reviewer File")

        with col2:
            FR = st.file_uploader('Analyst File')
    else:
        st.warning("Enter Reviewer name...To proceed...!")
    try:
        AR_file_name = str(AR.name).split("_")[0]
        FR_file_name = str(FR.name).split("_")[0]

        if AR_file_name == FR_file_name:
            pass
        else:
            st.error("Both File tickers are miss macthing... Please look into it...!")
    except:
        pass
    
    def Delta(AR_f,FR_f):
        with st.spinner("Loading Excel Files...."):
            #Loading excel and activating it
            AR_df = load_workbook(AR_f)
            FR_df = load_workbook(FR_f)
            AR_fn = AR_df.active
            FR_fn = FR_df.active

        with st.spinner("Checking Columns...."):
            if AR_fn.max_column == FR_fn.max_column:
                with col6:
                    st_lottie(loader,height=200,width=200, key='loader')
            
            else:
                with col6:
                    st_lottie(error_gif,height=125,width=125,key='error_gif')
                st.error(f"Both files number of Columns are not same.! , Level 1 columns- {FR_fn.max_column}, Level 2 columns- {AR_fn.max_column}")
                st.stop()

        excel_files = [AR,FR]

        with st.spinner("Combinding Both Excels...."):
            # Create a new Excel workbook to consolidate the sheets.
            combined_workbook = openpyxl.Workbook()

            # Iterate through each Excel file and each sheet within each file.
            for excel_file in excel_files:
                workbook = openpyxl.load_workbook(excel_file)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Create a new sheet in the combined workbook with the same name.
                    combined_sheet = combined_workbook.create_sheet(title=sheet_name)
                    
                    # Copy data from the original sheet to the combined sheet.
                    for row in sheet.iter_rows():
                        for cell in row:
                            combined_sheet[cell.coordinate] = cell.value

            # Remove the default sheet created by openpyxl.
            combined_workbook.remove(combined_workbook.active)

            # Save the combined workbook to a new file.
            combined_workbook.save("combined_excel.xlsx")

        with st.spinner("Loading Combined file and Formating it...."):
            combined_wb = load_workbook("combined_excel.xlsx")
            AR_sheet = combined_wb['Sheet1']
            FR_sheet = combined_wb['Sheet11']

            #created new sheet to enter count
            combined_wb.create_sheet(title="Delta")

            #changing Sheet names
            AR_sheet.title = "Reviewer"
            FR_sheet.title = "Analyst"

            delta_sheet = combined_wb['Delta']

            #formating 
            font = Font(bold=True)
            border = Border(top=Side(style='thin'))

            #entering names in new sheet
            delta_sheet.cell(1,1).value='Errors'
            delta_sheet.cell(1,2).value='Count'
            delta_sheet.cell(2,1).value='Data Deleted'
            delta_sheet.cell(3,1).value='Data Added'
            delta_sheet.cell(4,1).value='Unit Error'
            delta_sheet.cell(5,1).value='Period Error'
            delta_sheet.cell(6,1).value='Merging Error'
            delta_sheet.cell(7,1).value= 'Wrong Tagging - Quater'
            delta_sheet.cell(8,1).value= 'Wrong Tagging - Value'
            delta_sheet.cell(9,1).value= 'Wrong Tagging - Data Replaced'
            delta_sheet.cell(10,1).value= 'Wrong Fiscal'
            delta_sheet.cell(11,1).value= 'Total Errors'

            #coloring
            delta_sheet.cell(3,3).fill=PatternFill(start_color="00FF00",fill_type="solid")
            delta_sheet.cell(4,3).fill=PatternFill(start_color="E6E220",fill_type="solid")
            delta_sheet.cell(5,3).fill=PatternFill(start_color="E69720",fill_type="solid")
            delta_sheet.cell(6,3).fill=PatternFill(start_color="FF0000",fill_type="solid")
            delta_sheet.cell(7,3).fill=PatternFill(start_color="0000FF",fill_type="solid")
            delta_sheet.cell(8,3).fill=PatternFill(start_color="0000FF",fill_type="solid")
            delta_sheet.cell(9,3).fill= PatternFill(start_color="0000FF",fill_type="solid")
            delta_sheet.cell(10,3).fill=PatternFill(start_color="D518DB",fill_type="solid")

            #formating this cell
            delta_sheet.cell(11,1).font = font
            delta_sheet.cell(11,1).border = border

        with st.spinner("Extracting SRC values...."):
            AR_src = extract_hyperlinks_from_excel(AR_f)
            FR_src = extract_hyperlinks_from_excel(FR_f)

            deleted_src = [item for item in FR_src if item not in AR_src]
            data_added_src = [item for item in AR_src if item not in FR_src]

        with st.spinner("Checking Fiscal dates...."):        
            #for fical dates
            fiscal_count = 0
            fiscal_count_cells = []
            for row in range(1, 4):
                for column in range (1, max(AR_fn.max_column, FR_fn.max_column) + 1):
                    cell1 = AR_fn.cell(row = row, column= column)
                    cell2 = FR_fn.cell(row = row, column=column)
                    if cell1.value == cell2.value:
                        pass
                    else:
                        note = f"Value Changed from {cell2.value}, to {cell1.value}"
                        cell = AR_sheet.cell(row,column)
                        fiscal_count_cells.append(cell.coordinate)
                        cell.fill = PatternFill(start_color="D518DB",fill_type="solid")
                        cell.comment = Comment(note, author="R. Praveen")
                        fiscal_count = fiscal_count+1
                        print(f"wrong fiscal -{cell1.value}, from-{cell2.value}")

        
        with st.spinner("Checking For Data points added & deleted...."):
            for row in range(1,FR_fn.max_row + 1):
                for column in range (1,FR_fn.max_column + 1):
                    cell = FR_fn.cell(row = row, column= column)
                    if cell.hyperlink is not None:
                        if (cell.hyperlink.target.split("/")[-1]) in deleted_src:
                            ro = cell.row
                            col= cell.column
                            cell_col = FR_sheet.cell(ro,col)
                            cell_col.fill = PatternFill(start_color="FF0000",fill_type="solid")
                            cell_col.comment = Comment("Data deleted in AR file", author="R. Praveen")

            for row in range(1,AR_fn.max_row + 1):
                for column in range (1,AR_fn.max_column + 1):
                    cell = AR_fn.cell(row = row, column= column)
                    if cell.hyperlink is not None:
                        if (cell.hyperlink.target.split("/")[-1]) in data_added_src:
                            ro = cell.row
                            col= cell.column
                            cell_col = AR_sheet.cell(ro,col)
                            #print(cell_col.coordinate)
                            cell_col.fill = PatternFill(start_color="00FF00",fill_type="solid")
                            cell_col.comment = Comment("Data Added in AR file", author="R. Praveen")

        with st.spinner("Checking For Unit & Period Errors...."):
            UP_dict_ar = unit_period_dict(AR_fn,deleted_src,data_added_src,AR_src)
            UP_dict_fr = unit_period_dict(FR_fn,deleted_src,data_added_src,AR_src)

            unit_count=[]
            unit_error_cells = []
            for item in UP_dict_fr:
                if item in UP_dict_ar:
                    if UP_dict_fr[item][item][0] == UP_dict_ar[item][item][0]:
                        pass
                    else:
                        ro=(UP_dict_ar[item][item][3]).row
                        cell_col = AR_sheet.cell(ro,3)
                        list_unit,dumy_ar,dumy_fr = All_SRC_in_ROW(AR_fn,ro,data_added_src,deleted_src)
                        unit_count.append(int(len(list_unit)))
                        cell_col.fill = PatternFill(start_color="E6E220",fill_type="solid")
                        unit_error_cells.append(cell_col.coordinate) # saving cell address
                        note = f'Unit is Changed from {UP_dict_fr[item][item][0]} to {UP_dict_ar[item][item][0]}'
                        cell_col.comment = Comment(note, author="R. Praveen")
                        #print(f'Unit changed in FR file at {UP_dict_fr[item][item][0]}, & in AR file {UP_dict_ar[item][item][0]}')

            #print(f"list of unit - {sum(unit_count)}")
            #print(unit_count)
            delta_sheet.cell(4,2).value = int(sum(unit_count))

            period_count = []
            period_error_cells = []
            for item in UP_dict_fr:
                if item in UP_dict_ar:
                    if UP_dict_fr[item][item][1] == UP_dict_ar[item][item][1]:
                        pass
                    else:
                        ro=(UP_dict_ar[item][item][3]).row
                        cell_col = AR_sheet.cell(ro,4)
                        list_unit_p,dumy_ar,dumy_fr = All_SRC_in_ROW(AR_fn,ro,data_added_src,deleted_src)
                        period_count.append(int(len(list_unit_p)))
                        cell_col.fill = PatternFill(start_color="E69720",fill_type="solid")
                        period_error_cells.append(cell_col.coordinate) #saving cells
                        note = f'Period is Changed from {UP_dict_fr[item][item][1]} to {UP_dict_ar[item][item][1]}'
                        cell_col.comment = Comment(note, author="R. Praveen")
                        #print(f'Period changed in FR file at {UP_dict_fr[item][item][1]}, & in AR file {UP_dict_ar[item][item][1]}')

            #print(f"list of period - {sum(period_count)}")
            #print(period_count)
            delta_sheet.cell(5,2).value = int(sum(period_count))

        with st.spinner("Per- processing for Merging & Wrong tagging...."):
            #Merging
            MER_ar = merge_unmerg_dict(AR_fn)
            MER_fr = merge_unmerg_dict(FR_fn)

            row_vise_src_FR = {}
            row_vise_src_AR = {}
            row_ar_wrong_tag = {}
            row_fr_wrong_tag = {}
            for item in MER_fr:
                if item in MER_ar:
                    row_list_fr = []
                    row_list_ar = []
                    row_fr = MER_fr[item][item][0]
                    row_ar = MER_ar[item][item][0]
                    row_list_fr,fr_,fr_wrong_tagging = All_SRC_in_ROW(FR_fn,row_fr,data_added_src,deleted_src)
                    row_list_ar,ar_wrong_tagging,ar_ = All_SRC_in_ROW(AR_fn,row_ar,data_added_src,deleted_src)
                    row_vise_src_FR[item] = row_list_fr
                    row_vise_src_AR[item] = row_list_ar
                    row_ar_wrong_tag[item] = ar_wrong_tagging
                    row_fr_wrong_tag[item] = fr_wrong_tagging

        with st.spinner("Checking Merging Errors...."):
            #list for merging count - Upgraded
            Merging_count = []
            unique_row_src = []
            for item in row_vise_src_FR:
                if item in row_vise_src_AR:
                    if row_vise_src_FR[item] == row_vise_src_AR[item]:
                        pass
                    else:
                        row = MER_fr[item][item][0]
                        row_ar = MER_ar[item][item][0]
                        row_c = AR_sheet[row_ar]
                        row_f_c = FR_sheet[row]
                        for cell in row_c:
                            cell.fill = PatternFill(start_color="FF0000",fill_type="solid")
                            note = f'Merging Error was corrected in this row'
                            cell.comment = Comment(note, author="R. Praveen")
                        for cell in row_f_c:
                            cell.fill = PatternFill(start_color="FF0000",fill_type="solid")
                            note = f'Merging Error was corrected in this row'
                            cell.comment = Comment(note, author="R. Praveen")
                        ar_count=len(row_vise_src_AR[item])
                        fr_count = len(row_vise_src_FR[item])
                        unique_elements = list(set(row_vise_src_FR[item]) ^ set(row_vise_src_AR[item]))
                        fr_row_num = []
                        for i in unique_elements:
                            fr_row_num.append(FR_src[i][i][-1].row)
                        fr_row_num = list(set(fr_row_num))
                        #print(len(fr_row_num))
                        if len(fr_row_num)<2:
                            if fr_row_num[0] in unique_row_src:
                                pass

                            else:
                                unique_row_src.append(fr_row_num[0])
                                print(fr_row_num)
                                for i in fr_row_num:
                                    final_count=0
                                    row_list_fr,fr_,fr_wrong_tagging = All_SRC_in_ROW(FR_fn,i,data_added_src,deleted_src)
                                    #print(f'{row_list_fr} --- {unique_elements}')
                                    if len(row_list_fr) == len(unique_elements): # Unmerged
                                        final_count = int(ar_count)
                                

                                    elif len(row_list_fr) >= len(unique_elements):
                                        final_count = fr_count

                                    print(f"""in FR file row no- {row}, was changed in AR file.- {fr_count}
                                        Row in AR file {row_ar}. - {ar_count}, final count -{final_count}""")
                                    #print(unique_elements)
                                    Merging_count.append(final_count)

                        else:
                            final_count=0
                            temp_count=0
                            for j in fr_row_num:
                                if j not in unique_row_src:
                                    unique_row_src.append(j)
                                    row_list_fr,fr_,fr_wrong_tagging = All_SRC_in_ROW(FR_fn,j,data_added_src,deleted_src)
                                    if len(row_list_fr) <= len(unique_elements): # Unmerged
                                        temp_count = int(ar_count)

                                    else:
                                        temp_count = temp_count + len(row_list_fr)
                        
                            final_count = temp_count
                            print(f"""in FR file row no- {row}, was changed in AR file.- {fr_count}
                                    Row in AR file {row_ar}. - {ar_count}, final count -{final_count}""")
                                #print(unique_elements)
                            Merging_count.append(final_count)


            

            delta_sheet.cell(6,2).value=int(sum(Merging_count))

        with st.spinner("Checking for Wrong Tagging - Quaters...."):
            # Wrog tagging - quater
            wrong_quater_tagged = []
            wrong_tagging_quater_cell=[]
            for item in FR_src.keys():
                if item in AR_src.keys():
                    if FR_src[item][item][2]==AR_src[item][item][2]:
                        pass
                    else:
                        #print(f"wrong taging in {FR_src[item][item][1]}, shifted to {AR_src[item][item][1]}")
                        fr_cell = FR_sheet.cell(FR_src[item][item][4].row,FR_src[item][item][4].column)
                        ar_cell = AR_sheet.cell(AR_src[item][item][4].row,AR_src[item][item][4].column)
                        ar_note = f'Wrong tagging corrected, shfited from {FR_src[item][item][1]} to {AR_src[item][item][1]}'
                        fr_cell.fill = PatternFill(start_color="0000FF",fill_type="solid")
                        fr_cell.comment = Comment('Wrong tagging', author="R. Praveen")
                        ar_cell.fill = PatternFill(start_color="0000FF",fill_type="solid")
                        ar_cell.comment = Comment(ar_note, author="R. Praveen")
                        wrong_quater_tagged.append(AR_src[item])
                        wrong_tagging_quater_cell.append(ar_cell.coordinate)
        
        with st.spinner("Checking for Wrong Tagging - Value...."):
            #wrong tagging - value
            Wrong_value_tagged = []
            wrong_value_tagged_cells = []
            for item in FR_src.keys():
                if item in AR_src.keys():
                    if FR_src[item][item][3]==AR_src[item][item][3]:
                        pass
                    else:
                        #print(f"wrong taging in {FR_src[item][item][1]}- {FR_src[item][item][3]}, changed in {AR_src[item][item][1]} , to - {AR_src[item][item][3]}")
                        fr_cell = FR_sheet.cell(FR_src[item][item][4].row,FR_src[item][item][4].column)
                        ar_cell = AR_sheet.cell(AR_src[item][item][4].row,AR_src[item][item][4].column)
                        ar_note = f'Wrong tagging corrected, Value changed from {FR_src[item][item][3]} to {AR_src[item][item][3]}'
                        fr_cell.fill = PatternFill(start_color="0000FF",fill_type="solid")
                        fr_cell.comment = Comment('Wrong tagging', author="R. Praveen")
                        ar_cell.fill = PatternFill(start_color="0000FF",fill_type="solid")
                        ar_cell.comment = Comment(ar_note, author="R. Praveen")
                        wrong_value_tagged_cells.append(ar_cell.coordinate)
                        Wrong_value_tagged.append(AR_src[item])

            delta_sheet.cell(7,2).value= int(len(wrong_quater_tagged))
            delta_sheet.cell(8,2).value= int(len(Wrong_value_tagged))

        with st.spinner("Checking for Wrong Tagging - Data replaced...."):
            # Wrong tagging - Data points replaced
            wrong_taging_dict = {}
            wrong_taging_dict_cells = []
            for item in row_fr_wrong_tag.keys():
                if item in row_ar_wrong_tag.keys():
                    if row_fr_wrong_tag[item] == row_ar_wrong_tag[item]:
                        pass
                    else:
                        ar = {}
                        fr = {}
                        fr_row_deleted = []
                        ar_row_added = []
                        for row_iter in row_fr_wrong_tag[item]:
                            if row_iter in deleted_src:
                            # print(f'deleted {row_iter}')
                                fr_row_deleted.append(row_iter)
                                
                        for row_iter in row_ar_wrong_tag[item]:
                            if row_iter in data_added_src:
                                ar_row_added.append(row_iter)
                        ar['AR']= ar_row_added
                        fr['FR']=fr_row_deleted
                        #print(f"fr - {len(fr_row_deleted)}")
                        if len(fr_row_deleted) > 0:
                            wrong_taging_dict[(AR_src[item][item][4]).row] = [ar,fr]

            AR_replaced = []
            FR_replaced = []
            for item in wrong_taging_dict.keys():
                for i in wrong_taging_dict[item][0]['AR']:
                    for j in wrong_taging_dict[item][1]['FR']:
                        if AR_src[i][i][2] == FR_src[j][j][2]:
                            AR_replaced.append(i)
                            FR_replaced.append(j)
                            cell_col = AR_sheet.cell((AR_src[i][i][4]).row,(AR_src[i][i][4]).column)
                            cell_col.fill = PatternFill(start_color="0000FF",fill_type="solid")
                            wrong_taging_dict_cells.append(cell_col.coordinate)
                            note = f'Wrong tagging, Number replaced with- {AR_src[i][i][4].value}, from- {FR_src[j][j][4].value}'
                            cell_col.comment = Comment(note, author="R. Praveen")

            delta_sheet.cell(9,2).value = int(len(AR_replaced))
        
        with st.spinner("Adjusting data points in Data Added & Deleted...."):
            #removing the replaced number from deleted and added count
            for item in FR_replaced:
                deleted_src.remove(item)

            for item in AR_replaced:
                data_added_src.remove(item)

        with st.spinner("Storing data in Delta Sheet...."):
            delta_sheet.cell(2,2).value= int(len(deleted_src))
            delta_sheet.cell(3,2).value=int(len(data_added_src))
            #print(f"Fiscal - {fiscal_count}")
            delta_sheet.cell(10,2).value = fiscal_count
            delta_sheet.cell(11,2).value= "=SUM(B2:B10)"
            

            delta_sheet.cell(7,4).value = str(wrong_tagging_quater_cell)
            delta_sheet.cell(4,4).value = str(unit_error_cells)
            delta_sheet.cell(5,4).value = str(period_error_cells)
            delta_sheet.cell(8,4).value = str(wrong_value_tagged_cells)
            delta_sheet.cell(9,4).value = str(wrong_taging_dict_cells)
            delta_sheet.cell(10,4).value = str(fiscal_count_cells)

            #Totaling and addtional formating
            delta_sheet.cell(15,1).value = 'Name of Errors'
            delta_sheet.cell(16,1).value = 'Missing Error'
            delta_sheet.cell(17,1).value = 'Merging Error'
            delta_sheet.cell(18,1).value = 'Wrong Tagging Error'
            delta_sheet.cell(19,1).value = 'Unit Error'
            delta_sheet.cell(20,1).value = 'Period Error'
            delta_sheet.cell(21,1).value = 'Wrong Fiscal'
            delta_sheet.cell(22,1).value = 'Total Errors'

            #Delta Errors
            delta_sheet.cell(15,2).value = 'Delta Errors'
            delta_sheet.cell(16,2).value = '=B3'
            delta_sheet.cell(17,2).value = '=B6'
            delta_sheet.cell(18,2).value = '=SUM(B7:B9)'
            delta_sheet.cell(19,2).value = '=B4'
            delta_sheet.cell(20,2).value = '=B5'
            delta_sheet.cell(21,2).value = '=B10'
            delta_sheet.cell(22,2).value = '=SUM(B3:B10)'

            #Diff
            delta_sheet.cell(15,3).value = 'Zoho Errors'
            delta_sheet.cell(15,4).value = 'Difference'
            delta_sheet.cell(16,4).value = '=B16-C16'
            delta_sheet.cell(17,4).value = '=B17-C17'
            delta_sheet.cell(18,4).value = '=B18-C18'
            delta_sheet.cell(19,4).value = '=B19-C19'
            delta_sheet.cell(20,4).value = '=B20-C20'
            delta_sheet.cell(21,4).value = '=B21-C21'
            delta_sheet.cell(22,3).value = '=SUM(C16:C21)'
            delta_sheet.cell(22,4).value = '=SUM(D16:D21)'


            #formating this cell 
            delta_sheet.cell(11,2).font = font
            delta_sheet.cell(11,2).border = border
            delta_sheet.cell(22,1).font = font
            delta_sheet.cell(22,1).border = border
            delta_sheet.cell(22,2).font = font
            delta_sheet.cell(22,2).border = border
            delta_sheet.cell(22,3).font = font
            delta_sheet.cell(22,3).border = border
            delta_sheet.cell(22,4).font = font
            delta_sheet.cell(22,4).border = border

            combined_wb.save("combined_excel.xlsx")
            
    
    d_but = st.button("Delta Review")
    

    col5,col6 = st.columns(2)
    download = False


    
    if d_but:
        with st.spinner("Reviewing...."):
            Delta(AR,FR)   
        download = True


    data = 'combined_excel.xlsx'
    # Read the file content
    with open(data, 'rb') as file:
        file_content = file.read()
    try:
        file_n = f'{AR_file_name}_{R_name}_delta.xlsx'
    except:
        pass

    if download:
        st.download_button("Download file",data=file_content,file_name=file_n,mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

elif selected == "Earnings":

    #st.subheader("Project in WIP....!")
    if R_name:
        col1,col2 = st.columns(2)
        with col1:
            AR = st.file_uploader("Reviewer File",accept_multiple_files=True)

        with col2:
            FR = st.file_uploader('Analyst File',accept_multiple_files=True)
    else:
        st.warning("Enter Reviewer name...To proceed...!")
    try:
        AR_file_name = str(AR.name).split("_")[0]
        FR_file_name = str(FR.name).split("_")[0]

        if AR_file_name == FR_file_name:
            pass
        else:
            st.error("Both File tickers are miss macthing... Please look into it...!")
    except:
        pass

    def Delta_earning(AR_f,FR_f,ticker_name,count):
        with st.spinner("Loading Excel Files..."):
            t1 = time.time()
            #Loading excel and activating it
            AR_df = load_workbook(AR_f)
            FR_df = load_workbook(FR_f)
            AR_fn = AR_df.active
            FR_fn = FR_df.active

        with st.spinner("Checking Columns...."):
            if AR_fn.max_column == FR_fn.max_column:
                with col6:
                    pass 
                    #st_lottie(loader,height=200,width=200, key='loader')
            
            else:
                with col6:
                    st_lottie(error_gif,height=125,width=125,key='error_gif')
                st.error(f"Both files number of Columns are not same.! , Level 1 columns- {FR_fn.max_column}, Level 2 columns- {AR_fn.max_column}")
                st.stop()

        excel_files = [AR_f,FR_f]

        # finding calender
        calender_value = AR_fn.cell(1,AR_fn.max_column).value

        # finding its latest quater is FY or not
        if calender_value[-2:] == 'FY':
            column_earning = -2
        else:
            column_earning = -1
        t2 = time.time()
        if count == 1:
            with st.spinner("Combining both excel...."):
                # Create a new Excel workbook to consolidate the sheets.
                combined_workbook = openpyxl.Workbook()

                # Iterate through each Excel file and each sheet within each file.
                name_change_count=0
                for excel_file in excel_files:
                    workbook = openpyxl.load_workbook(excel_file)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        if name_change_count == 0:
                            # Create a new sheet in the combined workbook with the same name.
                            combined_sheet = combined_workbook.create_sheet(title="AR")
                            name_change_count=name_change_count+1

                        else:
                            # Create a new sheet in the combined workbook with the same name.
                            combined_sheet = combined_workbook.create_sheet(title="FR")
                        
                        # Copy data from the original sheet to the combined sheet.
                        for row in sheet.iter_rows():
                            for cell in row:
                                combined_sheet[cell.coordinate] = cell.value
                try:
                    # Remove the default sheet created by openpyxl.
                    combined_workbook.remove(combined_workbook.active)
                except:
                    pass

                
        else:
            with st.spinner("Combining both excel...."):
                # Create a new Excel workbook to consolidate the sheets.
                #combined_workbook = openpyxl.Workbook()
                combined_workbook = load_workbook("combined_excel.xlsx")

                # Iterate through each Excel file and each sheet within each file.
                name_change_count=0
                for excel_file in excel_files:
                    workbook = openpyxl.load_workbook(excel_file)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        if name_change_count == 0:
                            # Create a new sheet in the combined workbook with the same name.
                            combined_sheet = combined_workbook.create_sheet(title="AR")
                            name_change_count=name_change_count+1

                        else:
                            # Create a new sheet in the combined workbook with the same name.
                            combined_sheet = combined_workbook.create_sheet(title="FR")
                        
                        # Copy data from the original sheet to the combined sheet.
                        for row in sheet.iter_rows():
                            for cell in row:
                                combined_sheet[cell.coordinate] = cell.value
                
        t3 = time.time()
        with st.spinner("Loading Combined Excel and formating...."):
            #combined_wb = load_workbook("combined_excel.xlsx")
            AR_sheet = combined_workbook['AR']
            FR_sheet = combined_workbook['FR']

        with st.spinner("Loading Combined Excel and formating...."):    
            #created new sheet to enter count
            combined_workbook.create_sheet(title=f'{ticker_name}_Delta')

            #changing Sheet names
            AR_sheet.title = f"{ticker_name}_Reviewer"
            FR_sheet.title = f"{ticker_name}_Analyst"

            delta_sheet = combined_workbook[f'{ticker_name}_Delta']

            #formating 
            font = Font(bold=True)
            border = Border(top=Side(style='thin'))

            #entering names in new sheet
            delta_sheet.cell(1,1).value='Errors'
            delta_sheet.cell(1,2).value='Count'
            delta_sheet.cell(2,1).value='Data Deleted'
            delta_sheet.cell(3,1).value='Data Added'
            delta_sheet.cell(4,1).value='Unit Error'
            delta_sheet.cell(5,1).value='Period Error'
            delta_sheet.cell(6,1).value='Merging Error'
            delta_sheet.cell(7,1).value= 'Wrong Tagging - Quater'
            delta_sheet.cell(8,1).value= 'Wrong Tagging - Value'
            delta_sheet.cell(9,1).value= 'Wrong Tagging - Data Replaced'
            delta_sheet.cell(10,1).value= 'Wrong Fiscal'
            delta_sheet.cell(11,1).value= 'Total Errors'

            #coloring
            delta_sheet.cell(3,3).fill=PatternFill(start_color="00FF00",fill_type="solid")
            delta_sheet.cell(4,3).fill=PatternFill(start_color="E6E220",fill_type="solid")
            delta_sheet.cell(5,3).fill=PatternFill(start_color="E69720",fill_type="solid")
            delta_sheet.cell(6,3).fill=PatternFill(start_color="FF0000",fill_type="solid")
            delta_sheet.cell(7,3).fill=PatternFill(start_color="0000FF",fill_type="solid")
            delta_sheet.cell(8,3).fill=PatternFill(start_color="0000FF",fill_type="solid")
            delta_sheet.cell(9,3).fill= PatternFill(start_color="0000FF",fill_type="solid")
            delta_sheet.cell(10,3).fill=PatternFill(start_color="D518DB",fill_type="solid")

            #formating this cell
            delta_sheet.cell(11,1).font = font
            delta_sheet.cell(11,1).border = border

        t4 = time.time()
        with st.spinner("Extracting All SRC values...."):
            fr_all_src = extract_hyperlinks_from_excel_earnings(FR_f,column_earning)
            ar_all_src = extract_hyperlinks_from_excel_earnings(AR_f,column_earning)

            deleted_src = [item for item in fr_all_src if item not in ar_all_src]
            data_added_src = [item for item in ar_all_src if item not in fr_all_src]

        t5 = time.time()
        with st.spinner("Checking Fiscal dates...."):
            #for fical dates
            fiscal_count = 0
            fiscal_count_cells = []
            for row in range(1, 4):
                for column in range (1, max(AR_fn.max_column, FR_fn.max_column) + 1):
                    cell1 = AR_fn.cell(row = row, column= column)
                    cell2 = FR_fn.cell(row = row, column=column)
                    if cell1.value == cell2.value:
                        pass
                    else:
                        note = f"Value Changed from {cell2.value}, to {cell1.value}"
                        cell = AR_sheet.cell(row,column)
                        fiscal_count_cells.append(cell.coordinate)
                        cell.fill = PatternFill(start_color="D518DB",fill_type="solid")
                        cell.comment = Comment(note, author="R. Praveen")
                        fiscal_count = fiscal_count+1
                        print(f"wrong fiscal -{cell1.value}, from-{cell2.value}")

        t6 = time.time()
        with st.spinner("Checking Unit Errors...."):
            #Unit error
            unit_count=[]
            unit_cell_coordinate = []
            for item in fr_all_src:
                if item in ar_all_src:
                    if fr_all_src[item][item][6] == ar_all_src[item][item][6]:
                        pass
                    else:
                        print(f"unit was changed in AR cell - {ar_all_src[item][item][4].coordinate}")
                        cell_col = AR_sheet.cell(ar_all_src[item][item][4].row,ar_all_src[item][item][4].column)
                        cell_col.fill = PatternFill(start_color="E6E220",fill_type="solid")
                        unit_count.append(cell_col.coordinate) # saving cell address
                        note = f'Unit is Changed from {fr_all_src[item][item][6]} to {ar_all_src[item][item][6]}'
                        cell_col.comment = Comment(note, author="R. Praveen")
                        unit_cell_coordinate.append(ar_all_src[item][item][4].coordinate)
            
            delta_sheet.cell(4,2).value = int(len(unit_count))

        t7 = time.time()
        with st.spinner("Checking Period Errors...."):
            #Period error
            period_count = []
            period_cell_coordinate = []
            for item in fr_all_src:
                if item in ar_all_src:
                    if fr_all_src[item][item][7] == ar_all_src[item][item][7]:
                        pass
                    else:
                        print(f"period was changed in AR cell - {ar_all_src[item][item][4].coordinate}")
                        cell_col = AR_sheet.cell(ar_all_src[item][item][4].row,ar_all_src[item][item][4].column)
                        cell_col.fill = PatternFill(start_color="E69720",fill_type="solid")
                        period_count.append(cell_col.coordinate) #saving cells
                        note = f'Period is Changed from {fr_all_src[item][item][7]} to {ar_all_src[item][item][7]}'
                        cell_col.comment = Comment(note, author="R. Praveen")
                        period_cell_coordinate.append(ar_all_src[item][item][4].coordinate)

            delta_sheet.cell(5,2).value = int(len(period_count))

        t8 = time.time()
        with st.spinner("Checking Wrong Tagging - Quaters...."):
            #Wrong tagging
    # -----Wrong quater tagging-----
            wrong_quater_tagged = []
            wrong_quater_tagged_cell_coordinate = []
            for item in fr_all_src:
                if item in ar_all_src:
                    if fr_all_src[item][item][2] == ar_all_src[item][item][2]:
                        pass
                    else:
                        print(f"Wrong quater tagging, shifted to - {ar_all_src[item][item][4].coordinate}")
                        ar_cell = AR_sheet.cell(ar_all_src[item][item][4].row,ar_all_src[item][item][4].column)
                        ar_note = f'Wrong tagging corrected, shfited from {fr_all_src[item][item][1]} to {ar_all_src[item][item][1]}'
                        ar_cell.fill = PatternFill(start_color="0000FF",fill_type="solid")
                        ar_cell.comment = Comment(ar_note, author="R. Praveen")
                        wrong_quater_tagged.append(item)
                        wrong_quater_tagged_cell_coordinate.append(ar_all_src[item][item][4].coordinate)

        t9 = time.time()
        with st.spinner("Checking Wrong Tagging - Values...."):
    #------ Wrong value tagged ------
            Wrong_value_tagged = []
            wrong_value_tagged_cell_coordinate = []
            for item in fr_all_src:
                if item in ar_all_src:
                    if fr_all_src[item][item][3] == ar_all_src[item][item][3]:
                        pass
                    else:
                        print(f"Wrong value tagged , chaged in AR file - {ar_all_src[item][item][4].coordinate}")
                        ar_cell = AR_sheet.cell(ar_all_src[item][item][4].row,ar_all_src[item][item][4].column)
                        ar_note = f'Wrong tagging corrected, Value changed from {fr_all_src[item][item][3]} to {ar_all_src[item][item][3]}'
                        ar_cell.fill = PatternFill(start_color="0000FF",fill_type="solid")
                        ar_cell.comment = Comment(ar_note, author="R. Praveen")
                        Wrong_value_tagged.append(item)
                        wrong_value_tagged_cell_coordinate.append(ar_all_src[item][item][4].coordinate)

            delta_sheet.cell(7,2).value= int(len(wrong_quater_tagged))
            delta_sheet.cell(8,2).value= int(len(Wrong_value_tagged))

        t10 = time.time()
        with st.spinner("Collecting all SRC to check merging and data replaced...."):
            #Merging
            MER_ar = merge_unmerg_dict(AR_fn)
            MER_fr = merge_unmerg_dict(FR_fn)

            row_vise_src_FR = {}
            row_vise_src_AR = {}
            row_ar_wrong_tag = {}
            row_fr_wrong_tag = {}
            for item in MER_fr:
                if item in MER_ar:
                    row_list_fr = []
                    row_list_ar = []
                    row_fr = MER_fr[item][item][0]
                    row_ar = MER_ar[item][item][0]
                    row_list_fr,fr_,fr_wrong_tagging = All_SRC_in_ROW(FR_fn,row_fr,data_added_src,deleted_src)
                    row_list_ar,ar_wrong_tagging,ar_ = All_SRC_in_ROW(AR_fn,row_ar,data_added_src,deleted_src)
                    row_vise_src_FR[item] = row_list_fr
                    row_vise_src_AR[item] = row_list_ar
                    row_ar_wrong_tag[item] = ar_wrong_tagging
                    row_fr_wrong_tag[item] = fr_wrong_tagging

        t11 = time.time()
        with st.spinner("Checking Merging Errors...."):
            merge_row_latest_src=[]
            merge_row_previous_src=[]
            #storing the src of latest and previous columns where merging error is there
            for item in row_vise_src_FR:
                if item in row_vise_src_AR:
                    if row_vise_src_FR[item] == row_vise_src_AR[item]:
                        pass
                    else:
                        merge_row_latest_src.append(item)
                        try:
                            merge_row_previous_src.append(row_vise_src_FR[item][-2])
                        except:
                            pass

            #for latest
            Merging_count = []
            for item in merge_row_latest_src:
                if item in ar_all_src:
                    if ar_all_src[item][item][5] == fr_all_src[item][item][5]:
                        pass
                    else:
                        print(f"Merging error at - {ar_all_src[item][item][4].coordinate}")
                        cell=AR_sheet.cell(ar_all_src[item][item][4].row,ar_all_src[item][item][4].column)
                        cell.fill = PatternFill(start_color="FF0000",fill_type="solid")
                        note = f'Merging Error was corrected. shifted from {fr_all_src[item][item][4].coordinate} to {ar_all_src[item][item][4].coordinate}'
                        cell.comment = Comment(note, author="R. Praveen")
                        Merging_count.append(item)

            for item in merge_row_previous_src:
                if item in ar_all_src:
                    if ar_all_src[item][item][5] == fr_all_src[item][item][5]:
                        pass
                    else:
                        print(f"Merging error at - {ar_all_src[item][item][4].coordinate}")
                        cell=AR_sheet.cell(ar_all_src[item][item][4].row,ar_all_src[item][item][4].column)
                        cell.fill = PatternFill(start_color="FF0000",fill_type="solid")
                        note = f'Merging Error was corrected. shifted from {fr_all_src[item][item][4].coordinate} to {ar_all_src[item][item][4].coordinate}'
                        cell.comment = Comment(note, author="R. Praveen")
                        Merging_count.append(item)

        t12 = time.time()
        with st.spinner("Checking Wrong Tagging - Data Replaced...."):
            # Wrong tagging - Data points replaced
            wrong_taging_dict = {}
            wrong_taging_dict_cells = []
            for item in row_fr_wrong_tag.keys():
                if item in row_ar_wrong_tag.keys():
                    if row_fr_wrong_tag[item] == row_ar_wrong_tag[item]:
                        pass
                    else:
                        ar = {}
                        fr = {}
                        fr_row_deleted = []
                        ar_row_added = []
                        for row_iter in row_fr_wrong_tag[item]:
                            if row_iter in deleted_src:
                            # print(f'deleted {row_iter}')
                                fr_row_deleted.append(row_iter)
                                
                        for row_iter in row_ar_wrong_tag[item]:
                            if row_iter in data_added_src:
                                ar_row_added.append(row_iter)
                        ar['AR']= ar_row_added
                        fr['FR']=fr_row_deleted
                        #print(f"fr - {len(fr_row_deleted)}")
                        if len(fr_row_deleted) > 0:
                            wrong_taging_dict[(ar_all_src[item][item][4]).row] = [ar,fr]

            AR_replaced = []
            FR_replaced = []
            for item in wrong_taging_dict.keys():
                for i in wrong_taging_dict[item][0]['AR']:
                    for j in wrong_taging_dict[item][1]['FR']:
                        if ar_all_src[i][i][2] == fr_all_src[j][j][2]:
                            AR_replaced.append(i)
                            FR_replaced.append(j)
                            cell_col = ar_all_src.cell((ar_all_src[i][i][4]).row,(ar_all_src[i][i][4]).column)
                            cell_col.fill = PatternFill(start_color="0000FF",fill_type="solid")
                            wrong_taging_dict_cells.append(cell_col.coordinate)
                            note = f'Wrong tagging, Number replaced with- {ar_all_src[i][i][4].value}, from- {fr_all_src[j][j][4].value}'
                            cell_col.comment = Comment(note, author="R. Praveen")


            delta_sheet.cell(6,2).value=int(len(Merging_count))
            delta_sheet.cell(9,2).value = int(len(AR_replaced))

        t13 = time.time()
        with st.spinner("Adjusting Data added and Data deleted...."):
            for item in FR_replaced:
                deleted_src.remove(item)

            for item in AR_replaced:
                data_added_src.remove(item)

        data_added_cell = []
        for src_dp in data_added_src:
            cell = ar_all_src[src_dp][src_dp][4]
            data_added_cell.append(cell)
            
        t14 = time.time()
        with st.spinner("Storing all Error counts to Delta Sheet...."):
            delta_sheet.cell(2,2).value= int(len(deleted_src))
            delta_sheet.cell(3,2).value=int(len(data_added_src))
            delta_sheet.cell(3,3).value=str(data_added_cell)
            #print(f"Fiscal - {fiscal_count}")
            delta_sheet.cell(10,2).value = fiscal_count
            delta_sheet.cell(11,2).value= "=SUM(B2:B10)"

            delta_sheet.cell(7,4).value = str(wrong_quater_tagged_cell_coordinate)
            delta_sheet.cell(4,4).value = str(unit_cell_coordinate)
            delta_sheet.cell(5,4).value = str(period_cell_coordinate)
            delta_sheet.cell(8,4).value = str(wrong_value_tagged_cell_coordinate)
            delta_sheet.cell(9,4).value = str(wrong_taging_dict_cells)
            delta_sheet.cell(10,4).value = str(fiscal_count_cells)

            #Totaling and addtional formating
            delta_sheet.cell(15,1).value = 'Name of Errors'
            delta_sheet.cell(16,1).value = 'Missing Error'
            delta_sheet.cell(17,1).value = 'Merging Error'
            delta_sheet.cell(18,1).value = 'Wrong Tagging Error'
            delta_sheet.cell(19,1).value = 'Unit Error'
            delta_sheet.cell(20,1).value = 'Period Error'
            delta_sheet.cell(21,1).value = 'Wrong Fiscal'
            delta_sheet.cell(22,1).value = 'Total Errors'

            #Delta Errors
            delta_sheet.cell(15,2).value = 'Delta Errors'
            delta_sheet.cell(16,2).value = '=B3'
            delta_sheet.cell(17,2).value = '=B6'
            delta_sheet.cell(18,2).value = '=SUM(B7:B9)'
            delta_sheet.cell(19,2).value = '=B4'
            delta_sheet.cell(20,2).value = '=B5'
            delta_sheet.cell(21,2).value = '=B10'
            delta_sheet.cell(22,2).value = '=SUM(B3:B10)'

            #Diff
            delta_sheet.cell(15,3).value = 'Zoho Errors'
            delta_sheet.cell(15,4).value = 'Difference'
            delta_sheet.cell(16,4).value = '=B16-C16'
            delta_sheet.cell(17,4).value = '=B17-C17'
            delta_sheet.cell(18,4).value = '=B18-C18'
            delta_sheet.cell(19,4).value = '=B19-C19'
            delta_sheet.cell(20,4).value = '=B20-C20'
            delta_sheet.cell(21,4).value = '=B21-C21'
            delta_sheet.cell(22,3).value = '=SUM(C16:C21)'
            delta_sheet.cell(22,4).value = '=SUM(D16:D21)'


            #formating this cell 
            delta_sheet.cell(11,2).font = font
            delta_sheet.cell(11,2).border = border
            delta_sheet.cell(22,1).font = font
            delta_sheet.cell(22,1).border = border
            delta_sheet.cell(22,2).font = font
            delta_sheet.cell(22,2).border = border
            delta_sheet.cell(22,3).font = font
            delta_sheet.cell(22,3).border = border
            delta_sheet.cell(22,4).font = font
            delta_sheet.cell(22,4).border = border

            t15 = time.time()
            #timing 
            delta_sheet.cell(25,1).value = 'Process'
            delta_sheet.cell(26,1).value = 'Loading Excel'
            delta_sheet.cell(27,1).value = 'Combining Excel'
            delta_sheet.cell(28,1).value = 'Loading & Formating excel'
            delta_sheet.cell(29,1).value = 'Extracting all SRC'
            delta_sheet.cell(30,1).value = 'checking Fiscal'
            delta_sheet.cell(31,1).value = 'Checking Units'
            delta_sheet.cell(32,1).value = 'Checking Periods'
            delta_sheet.cell(33,1).value = 'checking Quaters'
            delta_sheet.cell(34,1).value = 'Checking values'
            delta_sheet.cell(35,1).value = 'SRC for Merging & Replaced data'
            delta_sheet.cell(36,1).value = 'Checking Merging Errors'
            delta_sheet.cell(37,1).value = 'Checking Data replaced'
            delta_sheet.cell(38,1).value = 'Adjusting Data points'
            delta_sheet.cell(39,1).value = 'Storing Delta Data'
            delta_sheet.cell(40,1).value = 'Total Time Taken'
            delta_sheet.cell(41,1).value = 'File size'

            delta_sheet.cell(25,2).value = "Time taken"
            delta_sheet.cell(26,2).value = int(t2-t1)
            delta_sheet.cell(27,2).value = int(t3-t2)
            delta_sheet.cell(28,2).value = int(t4-t3)
            delta_sheet.cell(29,2).value = int(t5-t4)
            delta_sheet.cell(30,2).value = int(t6-t5)
            delta_sheet.cell(31,2).value = int(t7-t6)
            delta_sheet.cell(32,2).value = int(t8-t7)
            delta_sheet.cell(33,2).value = int(t9-t8)
            delta_sheet.cell(34,2).value = int(t10-t9)
            delta_sheet.cell(35,2).value = int(t11-t10)
            delta_sheet.cell(36,2).value = int(t12-t11)
            delta_sheet.cell(37,2).value = int(t13-t12)
            delta_sheet.cell(38,2).value = int(t14-t13)
            delta_sheet.cell(39,2).value = int(t15-t14)
            delta_sheet.cell(40,2).value = int(t15-t1)
            delta_sheet.cell(41,2).value = f"{(AR_f.size/1000)}KB"
            

            file_name = "combined_excel.xlsx"
            combined_workbook.save(file_name)
            #return combined_workbook
            
    
    d_but = st.button("Delta Review")
    

    col5,col6 = st.columns(2)
    download = False

    
    
    if d_but:
        C=1
        with col6:
            st_lottie(loader,height=200,width=200, key='loader')
            for AR_file in AR:
                for FR_file in FR:
                    if str(AR_file.name).split("_")[0] == str(FR_file.name).split("_")[0]:
                        ticker_name= str(AR_file.name).split("_")[0]
                        with st.spinner(f"Reviewing....{ticker_name}"):
                            file = Delta_earning(AR_file,FR_file,ticker_name,C)   
                            C=C+1

        
        #file.save('combined_excel.xlsx')                            #st.write(len(AR))
        download = True


    data = 'combined_excel.xlsx'
    # Read the file content
    with open(data, 'rb') as file:
        file_content = file.read()
    try:
        file_n = f'{R_name}_Earnings_delta.xlsx'
    except:
        pass

    if download:
        st.download_button("Download file",data=file_content,file_name=file_n,mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        



        
    

